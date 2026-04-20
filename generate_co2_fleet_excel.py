# -*- coding: utf-8 -*-
"""
CO₂ Filo Ceza Hesaplaması — Adım Adım Excel Dosyası
EU 2019/1242 ve EU 2017/2400 regülasyonlarına dayalı
"""
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Renkler ──
DARK_BLUE = "1e3a5f"
BLUE = "2563eb"
LIGHT_BLUE = "dbeafe"
VERY_LIGHT_BLUE = "eff6ff"
RED = "E30613"
LIGHT_RED = "fee2e2"
GREEN = "059669"
LIGHT_GREEN = "d1fae5"
AMBER = "d97706"
LIGHT_AMBER = "fef3c7"
GRAY = "f1f5f9"
WHITE = "ffffff"
DARK = "0f172a"
DARK2 = "1e293b"

# ── Stiller ──
title_font = Font(name="Calibri", size=18, bold=True, color=WHITE)
title_fill = PatternFill("solid", fgColor=DARK_BLUE)
h1_font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
h2_font = Font(name="Calibri", size=12, bold=True, color=BLUE)
h3_font = Font(name="Calibri", size=11, bold=True, color=DARK2)
body_font = Font(name="Calibri", size=10, color=DARK)
bold_font = Font(name="Calibri", size=10, bold=True, color=DARK)
formula_font = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
result_font = Font(name="Calibri", size=12, bold=True, color=GREEN)
penalty_font = Font(name="Calibri", size=12, bold=True, color=RED)
small_font = Font(name="Calibri", size=9, color="64748b")
header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)

header_fill = PatternFill("solid", fgColor=DARK2)
formula_fill = PatternFill("solid", fgColor=VERY_LIGHT_BLUE)
result_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
penalty_fill = PatternFill("solid", fgColor=LIGHT_RED)
gray_fill = PatternFill("solid", fgColor=GRAY)
amber_fill = PatternFill("solid", fgColor=LIGHT_AMBER)
light_blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)

thin_border = Border(
    left=Side(style="thin", color="e2e8f0"),
    right=Side(style="thin", color="e2e8f0"),
    top=Side(style="thin", color="e2e8f0"),
    bottom=Side(style="thin", color="e2e8f0"),
)
thick_bottom = Border(
    bottom=Side(style="medium", color=DARK_BLUE),
)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

number_fmt = '#,##0.00'
int_fmt = '#,##0'
pct_fmt = '0.00%'
euro_fmt = '#,##0 €'


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def write_data_row(ws, row, values, start_col=1, bold=False, fill=None, fmt=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=v)
        cell.font = bold_font if bold else body_font
        cell.alignment = center if isinstance(v, (int, float)) else left
        cell.border = thin_border
        if fill:
            cell.fill = fill
        if fmt and isinstance(v, (int, float)):
            cell.number_format = fmt


def write_title(ws, row, text, col_span=8):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws.row_dimensions[row].height = 40


def write_section_header(ws, row, text, col_span=8, font=None):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font or h1_font
    cell.alignment = left
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws.row_dimensions[row].height = 28
    for c in range(1, col_span + 1):
        ws.cell(row=row, column=c).border = Border(bottom=Side(style="medium", color=BLUE))


def write_note(ws, row, text, col_span=8):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = small_font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)


def write_formula_box(ws, row, text, col_span=8):
    for c in range(1, col_span + 1):
        ws.cell(row=row, column=c).fill = formula_fill
        ws.cell(row=row, column=c).border = thin_border
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = formula_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws.row_dimensions[row].height = 30


def write_result_box(ws, row, label, value, unit="", col_span=8, is_penalty=False):
    # Label
    cell_l = ws.cell(row=row, column=1, value=label)
    cell_l.font = bold_font
    cell_l.alignment = right_align
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    # Value
    cell_v = ws.cell(row=row, column=4, value=value)
    cell_v.font = penalty_font if is_penalty else result_font
    cell_v.alignment = center
    if isinstance(value, float):
        cell_v.number_format = number_fmt
    elif isinstance(value, int):
        cell_v.number_format = int_fmt
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    # Unit
    cell_u = ws.cell(row=row, column=6, value=unit)
    cell_u.font = bold_font
    cell_u.alignment = left
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=col_span)
    # Fill
    fill = penalty_fill if is_penalty else result_fill
    for c in range(1, col_span + 1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 28


# ═══════════════════════════════════════
# SHEET 1: Genel Bakış & Regülasyon
# ═══════════════════════════════════════
def build_overview_sheet(wb):
    ws = wb.active
    ws.title = "1. Genel Bakış"
    set_col_widths(ws, [4, 24, 20, 20, 20, 20, 20, 20])
    ws.sheet_properties.tabColor = DARK_BLUE

    r = 1
    write_title(ws, r, "CO₂ FİLO CEZA HESAPLAMASI — ADIM ADIM KILAVUZ")
    r += 2

    write_section_header(ws, r, "📋 Bu Dosya Ne Anlatıyor?")
    r += 1
    for text in [
        "Bu Excel dosyası, AB regülasyonlarına göre CO₂ filo ceza hesaplamasını adım adım gösterir.",
        "Her sayfa bir hesaplama adımını temsil eder ve formüller, açıklamalar ve sayısal örnekler içerir.",
        "Amaç: TEMSA'nın otobüs filosunun CO₂ emisyon performansını ölçmek ve olası cezayı hesaplamak.",
    ]:
        write_note(ws, r, f"  • {text}")
        r += 1
    r += 1

    write_section_header(ws, r, "📌 Hesaplama Adımları (Sayfa Sırası)")
    r += 1
    steps = [
        ["1. Genel Bakış", "Bu sayfa — Regülasyon çerçevesi ve hesaplama zinciri"],
        ["2. VECTO & CO₂ Temeli", "Bireysel araç CO₂ hesaplama yöntemi (BSFC, düzeltme faktörleri)"],
        ["3. Filo Verisi", "Filo araçları, varyantlar, adetler ve CO₂ değerleri"],
        ["4. Filo CO₂ Ortalaması", "Ağırlıklı filo ortalaması hesabı (Σ(N×CO₂) / Σ(N))"],
        ["5. Alt Grup Analizi", "P31SD / P32SD bazlı ayrıştırma"],
        ["6. Hedef Hesaplama", "AB 2019/1242 azaltım hedeflerinin uygulanması"],
        ["7. Ceza Hesaplama", "Aşım tespiti ve ceza (€) hesabı"],
        ["8. ZLEV & Senaryo", "Sıfır emisyon araç etkisi ve alternatif senaryolar"],
    ]
    write_header_row(ws, r, ["Sayfa", "İçerik"])
    r += 1
    for step in steps:
        write_data_row(ws, r, step, bold=step[0].startswith("7"))
        r += 1
    r += 1

    write_section_header(ws, r, "📜 Yasal Çerçeve")
    r += 1
    write_header_row(ws, r, ["Regülasyon", "Konu", "Yürürlük"])
    r += 1
    regs = [
        ["EU 2019/1242", "Ağır hizmet araçları CO₂ performans standartları", "2019 → hedefler 2025+"],
        ["EU 2017/2400", "VECTO simülasyon aracı — CO₂ ve yakıt tüketimi belirleme", "2017"],
        ["EU 2018/956", "CO₂ izleme ve raporlama yükümlülükleri", "2018"],
        ["EU 2024/1610", "Güncellenmiş hedefler: 2035 %45, 2040 %90", "2024"],
        ["EN 590", "Dizel yakıt standardı — emisyon faktörü kaynağı", "Sürekli"],
    ]
    for reg in regs:
        write_data_row(ws, r, reg)
        r += 1
    r += 1

    write_section_header(ws, r, "🔗 Hesaplama Zinciri")
    r += 1
    write_header_row(ws, r, ["Adım", "İşlem", "Girdi", "Çıktı"])
    r += 1
    chain = [
        ["1", "BSFC hesaplama", "Yakıt haritası (RPM, tork, FC)", "BSFC (g/kWh)"],
        ["2", "Çevrim bazlı FC", "BSFC + çevrim parametreleri", "FC (g/km)"],
        ["3", "Düzeltme faktörleri", "FC + WHTC, BF, CF", "FC_corrected (g/km)"],
        ["4", "CO₂'ye dönüştürme", "FC_corrected × 3.169", "CO₂ (g/km)"],
        ["5", "Ağırlıklı ortalama", "CO₂ × çevrim ağırlıkları", "CO₂_weighted (g/km)"],
        ["6", "Filo ortalaması", "CO₂_v × N_v", "CO₂_fleet (g/km)"],
        ["7", "Hedef hesaplama", "Referans + azaltım oranları", "CO₂_target (g/km)"],
        ["8", "Ceza", "CO₂_fleet − CO₂_target", "€ / araç"],
    ]
    for row_data in chain:
        write_data_row(ws, r, row_data, fill=gray_fill if int(row_data[0]) % 2 == 0 else None)
        r += 1

    ws.sheet_view.showGridLines = False
    return ws


# ═══════════════════════════════════════
# SHEET 2: VECTO & CO₂ Temeli
# ═══════════════════════════════════════
def build_vecto_sheet(wb):
    ws = wb.create_sheet("2. VECTO & CO₂ Temeli")
    set_col_widths(ws, [4, 22, 18, 16, 16, 18, 18, 18])
    ws.sheet_properties.tabColor = BLUE

    r = 1
    write_title(ws, r, "VECTO METODOLOJİSİ — BİREYSEL ARAÇ CO₂ HESAPLAMA")
    r += 2

    # === Ana Formül ===
    write_section_header(ws, r, "ADIM 1: Temel CO₂ Formülü")
    r += 1
    write_formula_box(ws, r, "CO₂ [g/km] = FC_corrected [g/km] × EF_diesel")
    r += 1
    write_note(ws, r, "Kaynak: EU 2017/2400, Annex III, Madde 3.2")
    r += 2

    write_header_row(ws, r, ["Parametre", "Sembol", "Değer", "Birim", "Açıklama"])
    r += 1
    params = [
        ["Düzeltilmiş yakıt tüketimi", "FC_corrected", "—", "g/km", "Düzeltme faktörleri uygulanmış yakıt tüketimi"],
        ["Dizel emisyon faktörü", "EF_diesel", 3.169, "g CO₂ / g yakıt", "EN 590 standardından — C₁₂H₂₃ yanma ürünü"],
    ]
    for p in params:
        write_data_row(ws, r, p)
        r += 1
    r += 1

    # === EF Detay ===
    write_section_header(ws, r, "ADIM 2: Emisyon Faktörü Nereden Gelir?", font=h2_font)
    r += 1
    write_formula_box(ws, r, "C₁₂H₂₃ + 17.75 O₂ → 12 CO₂ + 11.5 H₂O")
    r += 1
    write_note(ws, r, "Mol kütleleri: C₁₂H₂₃ = 167 g/mol, 12×CO₂ = 528 g/mol → EF = 528/167 = 3.162 ≈ 3.169 (EN 590)")
    r += 2

    # === Düzeltme Faktörleri ===
    write_section_header(ws, r, "ADIM 3: Düzeltme Faktörleri")
    r += 1
    write_formula_box(ws, r, "FC_corrected = FC_base × WHTC_cycle × BF_ColdHot × CF_RegPer × CF_NCV")
    r += 1
    write_note(ws, r, "Kaynak: EU 2017/2400, Annex III, Madde 5.1 — Her çarpım bir fiziksel düzeltme kaynağını temsil eder")
    r += 2

    write_header_row(ws, r, ["Faktör", "Sembol", "Urban", "Rural", "Motorway", "Açıklama"])
    r += 1
    factors = [
        ["WHTC Çevrim", "WHTC", 1.08, 1.01, 1.00, "Geçici yük düzeltmesi"],
        ["Soğuk/Sıcak", "BF_ColdHot", 1.005, 1.005, 1.005, "Soğuk çalıştırma etkisi (+%0.5)"],
        ["Periyodik", "CF_RegPer", 1.0, 1.0, 1.0, "Regülasyon periyodik düzeltme"],
        ["NCV", "CF_NCV", 0.994, 0.994, 0.994, "Net kalorifik değer düzeltmesi (−%0.6)"],
    ]
    for f_data in factors:
        write_data_row(ws, r, f_data, fmt=number_fmt)
        r += 1
    r += 1

    # Toplam düzeltme faktörleri
    write_section_header(ws, r, "Toplam Düzeltme Çarpanları", font=h3_font)
    r += 1
    totals = [
        ["Urban", "=", "1.08 × 1.005 × 1.0 × 0.994", 1.0785],
        ["Rural", "=", "1.01 × 1.005 × 1.0 × 0.994", 1.0089],
        ["Motorway", "=", "1.00 × 1.005 × 1.0 × 0.994", 0.9990],
    ]
    for t in totals:
        write_data_row(ws, r, t, bold=True, fill=formula_fill)
        r += 1
    r += 1

    # === Çevrim Parametreleri ===
    write_section_header(ws, r, "ADIM 4: Seyir Çevrimi Parametreleri")
    r += 1
    write_header_row(ws, r, ["Çevrim", "Ort. Hız (km/h)", "Güç Fraksiyonu", "BSFC Seçimi", "Açıklama"])
    r += 1
    cycles = [
        ["Urban (Şehir İçi)", 18, "35%", "P75 (en yüksek)", "Sık dur-kalk, düşük verim bölgesi"],
        ["Rural (Kırsal)", 45, "45%", "Ortalama", "Karma şartlar, orta verim"],
        ["Motorway (Otoyol)", 70, "55%", "P25 (en düşük)", "Sabit hız, optimum verim"],
    ]
    for c in cycles:
        write_data_row(ws, r, c)
        r += 1
    r += 1

    # === Sayısal Örnek ===
    write_section_header(ws, r, "ADIM 5: Sayısal Örnek — Tek Araç CO₂ Hesabı")
    r += 1
    write_note(ws, r, "Varsayımlar: Motor gücü = 250 kW, BSFC_urban = 240 g/kWh, BSFC_rural = 215 g/kWh, BSFC_motorway = 200 g/kWh")
    r += 2

    write_header_row(ws, r, ["Adım", "Çevrim", "Hesaplama", "Sonuç", "Birim"])
    r += 1
    calc_steps = [
        ["1. Ortalama güç", "Urban", "250 × 0.35", 87.5, "kW"],
        ["2. Saatlik FC", "Urban", "240 × 87.5", 21000, "g/h"],
        ["3. Km başına FC", "Urban", "21000 / 18", 1166.67, "g/km"],
        ["4. Düzeltilmiş FC", "Urban", "1166.67 × 1.0785", 1258.25, "g/km"],
        ["5. CO₂", "Urban", "1258.25 × 3.169", 3987.39, "g/km"],
        ["", "", "", "", ""],
        ["1. Ortalama güç", "Rural", "250 × 0.45", 112.5, "kW"],
        ["2. Saatlik FC", "Rural", "215 × 112.5", 24187.5, "g/h"],
        ["3. Km başına FC", "Rural", "24187.5 / 45", 537.50, "g/km"],
        ["4. Düzeltilmiş FC", "Rural", "537.50 × 1.0089", 542.29, "g/km"],
        ["5. CO₂", "Rural", "542.29 × 3.169", 1718.47, "g/km"],
        ["", "", "", "", ""],
        ["1. Ortalama güç", "Motorway", "250 × 0.55", 137.5, "kW"],
        ["2. Saatlik FC", "Motorway", "200 × 137.5", 27500, "g/h"],
        ["3. Km başına FC", "Motorway", "27500 / 70", 392.86, "g/km"],
        ["4. Düzeltilmiş FC", "Motorway", "392.86 × 0.9990", 392.47, "g/km"],
        ["5. CO₂", "Motorway", "392.47 × 3.169", 1243.71, "g/km"],
    ]
    for s in calc_steps:
        fill = None
        if s[0] == "5. CO₂":
            fill = result_fill
        write_data_row(ws, r, s, bold=(s[0].startswith("5.")), fill=fill)
        r += 1
    r += 1

    # === Ağırlıklı Ortalama ===
    write_section_header(ws, r, "ADIM 6: Ağırlıklı Ortalama CO₂")
    r += 1
    write_formula_box(ws, r, "CO₂_weighted = w_urban × CO₂_urban + w_rural × CO₂_rural + w_motorway × CO₂_motorway")
    r += 1
    write_note(ws, r, "Ağırlıklar: Urban=0.35, Rural=0.35, Motorway=0.30 — Toplamı 1.00")
    r += 2

    write_header_row(ws, r, ["Çevrim", "Ağırlık", "CO₂ (g/km)", "Katkı (g/km)"])
    r += 1
    weighted = [
        ["Urban", 0.35, 3987.39, 0.35 * 3987.39],
        ["Rural", 0.35, 1718.47, 0.35 * 1718.47],
        ["Motorway", 0.30, 1243.71, 0.30 * 1243.71],
    ]
    for w in weighted:
        write_data_row(ws, r, w, fmt=number_fmt)
        r += 1
    co2_weighted = sum(w[3] for w in weighted)
    write_result_box(ws, r, "CO₂_weighted =", round(co2_weighted, 2), "g/km")
    r += 1
    write_note(ws, r, "Bu değer, bu aracın filoya katılma değeridir (VECTO sertifika değeri ile aynı mantık)")

    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════
# SHEET 3: Filo Verisi
# ═══════════════════════════════════════
def build_fleet_data_sheet(wb):
    ws = wb.create_sheet("3. Filo Verisi")
    set_col_widths(ws, [4, 22, 18, 16, 16, 18, 16, 18])
    ws.sheet_properties.tabColor = "059669"

    r = 1
    write_title(ws, r, "FİLO VERİSİ — VARYANTLAR VE ARAÇ ADETLERİ")
    r += 2

    write_section_header(ws, r, "Filo Tanımı")
    r += 1
    write_note(ws, r, "Bu tabloda filodaki araç varyantları, CO₂ değerleri ve adetleri yer alır.")
    write_note(ws, r + 1, "Her varyantın CO₂ değeri VECTO simülasyonundan gelen sertifika değeridir.")
    r += 3

    write_header_row(ws, r, ["#", "Varyant", "Model", "Alt Grup", "CO₂ (g/km)", "FC (L/100km)", "Filo Adedi", "Tip"])
    r += 1
    variants = [
        [1, "LD SB E6-300", "LD SB", "P32SD", 812, 38.2, 80, "Dizel"],
        [2, "Avenue EV-250", "Avenue", "P31SD", 0, 0, 30, "Elektrikli (ZEV)"],
        [3, "Prestij CI-250", "Prestij", "P31SD", 925, 43.5, 50, "Dizel"],
        [4, "Maraton E6-350", "Maraton", "P32SD", 780, 36.7, 60, "Dizel"],
        [5, "NEO CI-200", "NEO", "P31SD", 890, 41.8, 40, "Dizel"],
    ]
    for v in variants:
        fill = light_blue_fill if v[7] == "Elektrikli (ZEV)" else None
        write_data_row(ws, r, v, fill=fill, fmt=number_fmt)
        r += 1

    # Toplam satırı
    total_count = sum(v[6] for v in variants)
    r += 1
    write_data_row(ws, r, ["", "", "", "TOPLAM", "", "", total_count, ""], bold=True, fill=amber_fill)
    r += 2

    # Alt Grup Özeti
    write_section_header(ws, r, "Alt Grup Bazlı Özet")
    r += 1
    write_header_row(ws, r, ["Alt Grup", "Tanım", "Araç Sayısı", "Varyant Sayısı"])
    r += 1
    sg_data = [
        ["P31SD", "Tek katlı şehir otobüsü (Class I/II)", 120, 3],
        ["P32SD", "Tek katlı şehirler arası (Class III)", 140, 2],
    ]
    for sg in sg_data:
        write_data_row(ws, r, sg)
        r += 1
    r += 1

    write_note(ws, r, "💡 EV (Avenue) zE (sıfır emisyon) araç olarak filoda — CO₂ = 0 g/km olarak hesaba girer")
    r += 1
    write_note(ws, r, "💡 N_v (filo adedi) = o varyanttan kaç araç üretilip satıldığı")

    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════
# SHEET 4: Filo CO₂ Ortalaması
# ═══════════════════════════════════════
def build_fleet_avg_sheet(wb):
    ws = wb.create_sheet("4. Filo CO₂ Ortalaması")
    set_col_widths(ws, [4, 22, 18, 18, 20, 18, 20, 18])
    ws.sheet_properties.tabColor = "d29922"

    r = 1
    write_title(ws, r, "FİLO CO₂ ORTALAMASI HESABI")
    r += 2

    write_section_header(ws, r, "Formül")
    r += 1
    write_formula_box(ws, r, "CO₂_fleet = Σ(N_v × CO₂_v) / Σ(N_v)")
    r += 1
    write_note(ws, r, "Kaynak: EU 2019/1242, Madde 4, §1 — Üreticinin ortalama spesifik CO₂ emisyonu")
    write_note(ws, r + 1, "N_v = varyant araç sayısı, CO₂_v = varyantın VECTO CO₂ değeri (g/km)")
    r += 3

    write_section_header(ws, r, "Adım Adım Hesaplama")
    r += 1
    write_header_row(ws, r, ["Varyant", "CO₂_v (g/km)", "N_v (adet)", "N_v × CO₂_v", "Katkı (%)", "Açıklama"])
    r += 1

    fleet_data = [
        ["LD SB E6-300", 812, 80],
        ["Avenue EV-250", 0, 30],
        ["Prestij CI-250", 925, 50],
        ["Maraton E6-350", 780, 60],
        ["NEO CI-200", 890, 40],
    ]
    total_n = sum(d[2] for d in fleet_data)
    total_nco2 = sum(d[1] * d[2] for d in fleet_data)

    for d in fleet_data:
        nco2 = d[1] * d[2]
        pct = (nco2 / total_nco2 * 100) if total_nco2 > 0 else 0
        desc = "Sıfır emisyon — filo ortalamasını düşürür" if d[1] == 0 else f"Filo ağırlığı: {d[2]}/{total_n}"
        fill = light_blue_fill if d[1] == 0 else None
        write_data_row(ws, r, [d[0], d[1], d[2], nco2, round(pct, 1), desc], fill=fill)
        r += 1

    # Toplam
    r += 1
    write_data_row(ws, r, ["TOPLAM", "", total_n, total_nco2, "100%", ""], bold=True, fill=amber_fill)
    r += 2

    # Sonuç
    fleet_avg = total_nco2 / total_n
    write_section_header(ws, r, "Sonuç")
    r += 1
    write_formula_box(ws, r, f"CO₂_fleet = {total_nco2:,} / {total_n} = {fleet_avg:.2f} g/km")
    r += 1
    write_result_box(ws, r, "Filo CO₂ Ortalaması =", round(fleet_avg, 2), "g/km")
    r += 2

    # Katkı analizi
    write_section_header(ws, r, "Katkı Analizi")
    r += 1
    write_note(ws, r, f"• Avenue EV (0 g/km, 30 adet): Filo ortalamasını {30 * fleet_avg / total_n:.1f} g/km düşürüyor")
    r += 1
    write_note(ws, r, f"• Prestij (925 g/km, 50 adet): En yüksek CO₂ katkısını yapıyor")
    r += 1
    write_note(ws, r, f"• Toplam {total_n} araçta ağırlıklı ortalama = {fleet_avg:.2f} g/km")

    ws.sheet_view.showGridLines = False
    return fleet_avg, total_n, total_nco2


# ═══════════════════════════════════════
# SHEET 5: Alt Grup Analizi
# ═══════════════════════════════════════
def build_subgroup_sheet(wb):
    ws = wb.create_sheet("5. Alt Grup Analizi")
    set_col_widths(ws, [4, 22, 18, 18, 20, 18, 20, 18])
    ws.sheet_properties.tabColor = "8b5cf6"

    r = 1
    write_title(ws, r, "ALT GRUP BAZLI FİLO ANALİZİ")
    r += 2

    write_section_header(ws, r, "Formül")
    r += 1
    write_formula_box(ws, r, "CO₂_fleet,sg = Σ(N_v∈sg × CO₂_v,sg) / Σ(N_v∈sg)")
    r += 1
    write_note(ws, r, "sg ∈ {P31SD, P32SD} — Her alt grup ayrı hesaplanır")
    r += 2

    # P31SD
    write_section_header(ws, r, "Alt Grup: P31SD (Tek Katlı Şehir Otobüsü)")
    r += 1
    write_header_row(ws, r, ["Varyant", "CO₂ (g/km)", "Adet (N_v)", "N×CO₂", "Tip"])
    r += 1
    p31_data = [
        ["Avenue EV-250", 0, 30, 0, "EV"],
        ["Prestij CI-250", 925, 50, 46250, "Dizel"],
        ["NEO CI-200", 890, 40, 35600, "Dizel"],
    ]
    for d in p31_data:
        fill = light_blue_fill if d[4] == "EV" else None
        write_data_row(ws, r, d, fill=fill)
        r += 1
    r += 1
    p31_total_n = 120
    p31_total_nco2 = 81850
    p31_avg = p31_total_nco2 / p31_total_n
    write_data_row(ws, r, ["TOPLAM P31SD", "", p31_total_n, p31_total_nco2, ""], bold=True, fill=amber_fill)
    r += 1
    write_formula_box(ws, r, f"CO₂_P31SD = {p31_total_nco2:,} / {p31_total_n} = {p31_avg:.2f} g/km")
    r += 1
    write_result_box(ws, r, "P31SD Filo Ort. =", round(p31_avg, 2), "g/km")
    r += 2

    # P32SD
    write_section_header(ws, r, "Alt Grup: P32SD (Tek Katlı Şehirler Arası)")
    r += 1
    write_header_row(ws, r, ["Varyant", "CO₂ (g/km)", "Adet (N_v)", "N×CO₂", "Tip"])
    r += 1
    p32_data = [
        ["LD SB E6-300", 812, 80, 64960, "Dizel"],
        ["Maraton E6-350", 780, 60, 46800, "Dizel"],
    ]
    for d in p32_data:
        write_data_row(ws, r, d)
        r += 1
    r += 1
    p32_total_n = 140
    p32_total_nco2 = 111760
    p32_avg = p32_total_nco2 / p32_total_n
    write_data_row(ws, r, ["TOPLAM P32SD", "", p32_total_n, p32_total_nco2, ""], bold=True, fill=amber_fill)
    r += 1
    write_formula_box(ws, r, f"CO₂_P32SD = {p32_total_nco2:,} / {p32_total_n} = {p32_avg:.2f} g/km")
    r += 1
    write_result_box(ws, r, "P32SD Filo Ort. =", round(p32_avg, 2), "g/km")
    r += 2

    # Karşılaştırma
    write_section_header(ws, r, "Alt Grup Karşılaştırması")
    r += 1
    write_header_row(ws, r, ["Alt Grup", "Araç Sayısı", "CO₂ Ort. (g/km)", "EV Etkisi"])
    r += 1
    write_data_row(ws, r, ["P31SD", p31_total_n, round(p31_avg, 2), "EV sayesinde düşük"], fill=result_fill)
    r += 1
    write_data_row(ws, r, ["P32SD", p32_total_n, round(p32_avg, 2), "Sadece dizel"])
    r += 2
    write_note(ws, r, f"💡 P31SD ortalaması ({p31_avg:.2f}) EV sayesinde P32SD'den ({p32_avg:.2f}) düşük")
    r += 1
    write_note(ws, r, "💡 AB regülasyonu her alt grubu ayrı değerlendirir — bir grupta iyi olmak diğerini kurtarmaz")

    ws.sheet_view.showGridLines = False
    return p31_avg, p32_avg, p31_total_n, p32_total_n


# ═══════════════════════════════════════
# SHEET 6: Hedef Hesaplama
# ═══════════════════════════════════════
def build_target_sheet(wb, p31_avg, p32_avg, p31_n, p32_n):
    ws = wb.create_sheet("6. Hedef Hesaplama")
    set_col_widths(ws, [4, 24, 18, 18, 18, 18, 18, 18])
    ws.sheet_properties.tabColor = "06b6d4"

    r = 1
    write_title(ws, r, "AB 2019/1242 — FİLO HEDEF DEĞERİ HESAPLAMA")
    r += 2

    # Azaltım hedefleri
    write_section_header(ws, r, "AB Azaltım Hedefleri")
    r += 1
    write_header_row(ws, r, ["Dönem", "Azaltım Oranı", "Hedef Yılı", "Yasal Dayanak"])
    r += 1
    targets = [
        ["2025–2029", "15%", 2025, "EU 2019/1242, Madde 1(a)"],
        ["2030–2034", "30%", 2030, "EU 2019/1242, Madde 1(b)"],
        ["2035–2039", "45% (revize)", 2035, "EU 2024/1610"],
        ["2040+", "90% (revize)", 2040, "EU 2024/1610"],
    ]
    for t in targets:
        fill = light_blue_fill if t[0] == "2025–2029" else None
        write_data_row(ws, r, t, fill=fill, bold=(t[0] == "2025–2029"))
        r += 1
    r += 1
    write_note(ws, r, "⬆️ Mavi satır: Şu an geçerli dönem (2025-2029, %15 azaltım)")
    r += 2

    # Hedef formülü
    write_section_header(ws, r, "Spesifik Emisyon Hedefi Formülü")
    r += 1
    write_formula_box(ws, r, "T_sg = CO₂_ref,sg × (1 − rf_sg)")
    r += 1
    write_note(ws, r, "T_sg: Alt grup hedefi, CO₂_ref: 2019 referans değeri, rf: Azaltım yüzdesi")
    r += 2

    # Referans değerler (varsayım)
    co2_ref_p31 = 850
    co2_ref_p32 = 900
    rf = 0.15  # 2025 dönemi

    write_section_header(ws, r, "Hesaplama (2025 Dönemi — %15 Azaltım)")
    r += 1
    write_header_row(ws, r, ["Alt Grup", "CO₂_ref (g/km)", "Azaltım (%)", "(1 − rf)", "Hedef T_sg (g/km)", "Fiili CO₂ (g/km)", "Durum"])
    r += 1

    t_p31 = co2_ref_p31 * (1 - rf)
    t_p32 = co2_ref_p32 * (1 - rf)
    status_p31 = "✅ Hedef altında" if p31_avg <= t_p31 else "❌ Hedef üstünde"
    status_p32 = "✅ Hedef altında" if p32_avg <= t_p32 else "❌ Hedef üstünde"

    write_data_row(ws, r, ["P31SD", co2_ref_p31, "15%", 0.85, t_p31, round(p31_avg, 2), status_p31],
                   fill=result_fill if p31_avg <= t_p31 else penalty_fill)
    r += 1
    write_data_row(ws, r, ["P32SD", co2_ref_p32, "15%", 0.85, t_p32, round(p32_avg, 2), status_p32],
                   fill=result_fill if p32_avg <= t_p32 else penalty_fill)
    r += 2

    # Üretici bazında hedef
    write_section_header(ws, r, "Üretici Bazında Toplam Hedef")
    r += 1
    write_formula_box(ws, r, "T_manufacturer = Σ(N_sg × T_sg) / Σ(N_sg)")
    r += 1
    total_n = p31_n + p32_n
    t_mfr = (p31_n * t_p31 + p32_n * t_p32) / total_n
    write_note(ws, r, f"T_mfr = ({p31_n} × {t_p31:.2f} + {p32_n} × {t_p32:.2f}) / {total_n}")
    r += 1
    t_mfr_numer = p31_n * t_p31 + p32_n * t_p32
    write_formula_box(ws, r, f"T_mfr = {t_mfr_numer:,.2f} / {total_n} = {t_mfr:.2f} g/km")
    r += 1
    write_result_box(ws, r, "Üretici Hedefi =", round(t_mfr, 2), "g/km")

    ws.sheet_view.showGridLines = False
    return t_mfr


# ═══════════════════════════════════════
# SHEET 7: Ceza Hesaplama
# ═══════════════════════════════════════
def build_penalty_sheet(wb, fleet_avg, t_mfr, total_n):
    ws = wb.create_sheet("7. Ceza Hesaplama")
    set_col_widths(ws, [4, 24, 18, 18, 18, 18, 18, 18])
    ws.sheet_properties.tabColor = RED

    r = 1
    write_title(ws, r, "AŞIM EMİSYON PRİMİ (CEZA) HESAPLAMASI")
    r += 2

    write_section_header(ws, r, "ADIM 1: Aşım Tespiti")
    r += 1
    write_formula_box(ws, r, "E_excess = CO₂_fleet − T_manufacturer")
    r += 1
    write_note(ws, r, "E_excess < 0 → aşım yok (surplus/kredi), E_excess > 0 → ceza uygulanır")
    r += 2

    write_header_row(ws, r, ["Parametre", "Değer", "Birim", "Kaynak"])
    r += 1
    write_data_row(ws, r, ["Filo CO₂ ortalaması", round(fleet_avg, 2), "g/km", "Sayfa 4'ten"])
    r += 1
    write_data_row(ws, r, ["Üretici hedefi", round(t_mfr, 2), "g/km", "Sayfa 6'dan"])
    r += 1
    write_data_row(ws, r, ["Toplam araç sayısı", total_n, "adet", "Sayfa 3'ten"])
    r += 2

    e_excess = fleet_avg - t_mfr

    write_formula_box(ws, r, f"E_excess = {fleet_avg:.2f} − {t_mfr:.2f} = {e_excess:.2f} g/km")
    r += 1

    if e_excess > 0:
        write_result_box(ws, r, "Aşım =", round(e_excess, 2), "g/km  → CEZA UYGULANIR ⚠️", is_penalty=True)
    else:
        write_result_box(ws, r, "Aşım =", round(e_excess, 2), "g/km  → CEZA YOK ✅")
    r += 2

    # Ceza formülü
    write_section_header(ws, r, "ADIM 2: Birim Ceza")
    r += 1
    write_header_row(ws, r, ["Dönem", "Birim Ceza", "Birim"])
    r += 1
    write_data_row(ws, r, ["2025–2039", "4,250", "€ / (g CO₂/tkm)"], bold=True, fill=light_blue_fill)
    r += 1
    write_data_row(ws, r, ["2040+", "6,800 (taslak)", "€ / (g CO₂/tkm)"])
    r += 2

    # tkm dönüşümü
    write_section_header(ws, r, "ADIM 3: tkm Dönüşümü")
    r += 1
    write_formula_box(ws, r, "CO₂ [g/tkm] = CO₂ [g/km] / Payload [ton]")
    r += 1
    payload = 8  # ton varsayım
    write_note(ws, r, f"Varsayım: Ortalama payload = {payload} ton (yolcu ağırlığı)")
    r += 2

    if e_excess > 0:
        e_excess_tkm = e_excess / payload
        write_formula_box(ws, r, f"E_excess [g/tkm] = {e_excess:.2f} / {payload} = {e_excess_tkm:.4f} g CO₂/tkm")
        r += 2

        # Ceza hesabı
        write_section_header(ws, r, "ADIM 4: Toplam Ceza Hesabı")
        r += 1
        write_formula_box(ws, r, "Ceza [€] = E_excess [g/tkm] × 4250 [€/(g/tkm)] × N_total [araç]")
        r += 2

        write_header_row(ws, r, ["Parametre", "Değer", "Birim"])
        r += 1
        write_data_row(ws, r, ["E_excess", round(e_excess_tkm, 4), "g CO₂/tkm"])
        r += 1
        write_data_row(ws, r, ["Birim ceza", 4250, "€/(g CO₂/tkm)"])
        r += 1
        write_data_row(ws, r, ["Toplam araç", total_n, "adet"])
        r += 2

        penalty = e_excess_tkm * 4250 * total_n
        write_formula_box(ws, r, f"Ceza = {e_excess_tkm:.4f} × 4,250 × {total_n} = {penalty:,.2f} €")
        r += 1
        write_result_box(ws, r, "TOPLAM CEZA =", round(penalty, 2), "€", is_penalty=True)
        r += 2

        # Araç başına ceza
        per_vehicle = penalty / total_n
        write_result_box(ws, r, "Araç Başına Ceza =", round(per_vehicle, 2), "€ / araç", is_penalty=True)
        r += 2

        write_note(ws, r, f"⚠️ Toplam {penalty:,.0f} € ceza — bu miktarı azaltmak için EV oranı artırılabilir veya dizel varyantların CO₂ değerleri düşürülebilir")
    else:
        write_result_box(ws, r, "CEZA =", 0, "€  — Filo hedefin altında, ceza yok ✅")
        r += 2
        write_note(ws, r, f"✅ Surplus: {abs(e_excess):.2f} g/km hedefin altında — CO₂ kredisi biriktirilebilir")

    ws.sheet_view.showGridLines = False
    return e_excess


# ═══════════════════════════════════════
# SHEET 8: ZLEV & Senaryolar
# ═══════════════════════════════════════
def build_scenario_sheet(wb, fleet_avg, t_mfr, total_n):
    ws = wb.create_sheet("8. ZLEV & Senaryo")
    set_col_widths(ws, [4, 22, 16, 16, 16, 18, 18, 22])
    ws.sheet_properties.tabColor = "8b5cf6"

    r = 1
    write_title(ws, r, "ZLEV DÜZELTME & SENARYO ANALİZİ")
    r += 2

    # ZLEV Açıklama
    write_section_header(ws, r, "ZLEV (Sıfır ve Düşük Emisyonlu Araç) Faktörü")
    r += 1
    write_note(ws, r, "ZLEV: CO₂ emisyonu 0–1 g CO₂/kWh olan araçlar (tam elektrikli, yakıt hücreli, plug-in hibrid)")
    r += 2

    write_formula_box(ws, r, "ZLEV_factor = 1 − (share_ZLEV − benchmark_ZLEV) × multiplier")
    r += 1
    write_note(ws, r, "Kaynak: EU 2019/1242, Madde 5, §1")
    r += 2

    write_header_row(ws, r, ["Parametre", "2025–2029 Değeri", "Açıklama"])
    r += 1
    zlev_params = [
        ["share_ZLEV", f"{30/260*100:.1f}%  (30/260)", "Filodaki EV oranı"],
        ["benchmark_ZLEV", "2%", "AB minimum ZLEV benchmark"],
        ["multiplier", "0.03", "Esneklik çarpanı"],
        ["Maks. esneklik", "3%", "Hedef en fazla %3 esnetilebilir"],
    ]
    for p in zlev_params:
        write_data_row(ws, r, p)
        r += 1
    r += 1

    share_zlev = 30 / 260
    benchmark = 0.02
    multiplier = 0.03
    zlev_factor = 1 - (share_zlev - benchmark) * multiplier

    write_section_header(ws, r, "ZLEV Hesaplama")
    r += 1
    write_formula_box(ws, r, f"ZLEV_factor = 1 − ({share_zlev:.4f} − 0.02) × 0.03 = {zlev_factor:.6f}")
    r += 1
    t_adjusted = t_mfr / zlev_factor
    write_formula_box(ws, r, f"T_adjusted = {t_mfr:.2f} / {zlev_factor:.6f} = {t_adjusted:.2f} g/km")
    r += 1
    benefit = t_adjusted - t_mfr
    write_result_box(ws, r, "ZLEV Esnekliği =", round(benefit, 2), f"g/km eklendi hedefe (yeni hedef: {t_adjusted:.2f})")
    r += 3

    # Senaryolar
    write_section_header(ws, r, "Senaryo Analizi: EV Sayısı Değiştirme Etkisi")
    r += 1
    write_note(ws, r, "Diğer varyantlar sabit kalırken farklı EV adetleri ile filo ortalaması nasıl değişir?")
    r += 2

    write_header_row(ws, r, ["Senaryo", "EV Adedi", "Toplam Araç", "CO₂ Fleet", "Hedef", "Aşım", "Ceza (€)", "Sonuç"])
    r += 1

    # Dizel toplam (EV hariç)
    diesel_nco2 = 80 * 812 + 50 * 925 + 60 * 780 + 40 * 890  # = 193610
    diesel_n = 80 + 50 + 60 + 40  # = 230

    scenarios = [0, 10, 20, 30, 40, 50, 70, 100]
    for ev_count in scenarios:
        total = diesel_n + ev_count
        fleet_co2 = diesel_nco2 / total
        # Basit hedef hesabı (toplam üretici hedefi buradaki verilerle)
        target = t_mfr  # aynı hedef varsayımı
        excess = fleet_co2 - target
        payload = 8
        if excess > 0:
            penalty = (excess / payload) * 4250 * total
            result = "❌ CEZA"
            fill = penalty_fill
        else:
            penalty = 0
            result = "✅ SURPLUS"
            fill = result_fill
        write_data_row(ws, r, [
            f"EV={ev_count}", ev_count, total,
            round(fleet_co2, 2), round(target, 2),
            round(excess, 2), round(penalty, 0), result
        ], fill=fill)
        r += 1

    r += 1
    write_note(ws, r, "💡 EV sayısı arttıkça filo ortalaması düşer ve ceza azalır/sıfırlanır")
    r += 1
    write_note(ws, r, "💡 Ceza sıfırlandıktan sonraki her EV, CO₂ kredisi olarak birikir")
    r += 2

    # Break-even analizi
    write_section_header(ws, r, "Break-Even Analizi: Ceza Sıfırlama Noktası")
    r += 1
    # Cezanın sıfır olduğu EV sayısını bul
    for ev in range(0, 500):
        total = diesel_n + ev
        co2 = diesel_nco2 / total
        if co2 <= t_mfr:
            break_even_ev = ev
            break
    else:
        break_even_ev = "500+"

    write_formula_box(ws, r, f"Break-even EV adedi ≈ {break_even_ev} araç")
    r += 1
    write_note(ws, r, f"Bu noktada filo ortalaması ≤ {t_mfr:.2f} g/km ve ceza = 0 €")
    r += 1
    write_note(ws, r, f"Mevcut durum: {30} EV araç — {'ceza var' if fleet_avg > t_mfr else 'ceza yok'}")

    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════
# MAIN
# ═══════════════════════════════════════
def main():
    wb = Workbook()

    # Sheet 1
    build_overview_sheet(wb)

    # Sheet 2
    build_vecto_sheet(wb)

    # Sheet 3
    build_fleet_data_sheet(wb)

    # Sheet 4
    fleet_avg, total_n, total_nco2 = build_fleet_avg_sheet(wb)

    # Sheet 5
    p31_avg, p32_avg, p31_n, p32_n = build_subgroup_sheet(wb)

    # Sheet 6
    t_mfr = build_target_sheet(wb, p31_avg, p32_avg, p31_n, p32_n)

    # Sheet 7
    e_excess = build_penalty_sheet(wb, fleet_avg, t_mfr, total_n)

    # Sheet 8
    build_scenario_sheet(wb, fleet_avg, t_mfr, total_n)

    # Freeze panes for all sheets
    for ws in wb.worksheets:
        ws.freeze_panes = "A3"
        ws.print_options.gridLines = False

    output_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "CO2_Filo_Ceza_Hesaplama_Adim_Adim.xlsx"
    )
    wb.save(output_path)
    print(f"Excel oluşturuldu: {output_path}")
    print(f"Sayfa sayısı: {len(wb.worksheets)}")


if __name__ == "__main__":
    main()
